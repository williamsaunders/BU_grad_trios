from IPython.core.display import display, HTML
display(HTML("<style>.container { width:95% !important; }</style>"))

from matplotlib_inline import backend_inline
backend_inline.set_matplotlib_formats('retina')

import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook

import ipywidgets as widgets
from IPython.display import display, clear_output
from ipywidgets import HBox, VBox
from IPython.display import HTML
import base64

pd.set_option('display.max_rows', None)

def find_nearest(array, value):
    array = np.asarray(array)
    idx = (np.abs(array - value)).argmin()
    return idx

def init(df):
    df['prior_matches'] = [ [] for _ in range(len(df)) ]
    df['current_match'] = [ [] for _ in range(len(df)) ]
    df['current_group'] = -1
    df['num_prior_matches'] = 0
    df['size_prev_match'] = 0 # whether were in a 3 or 4 group previously, or 0 if new

def complete_match(df):
    # move the current match to the end of the list of prior matches
    for index, row in df.iterrows():
        df.loc[index, 'prior_matches'].extend(row['current_match'])
        df.loc[index, 'num_prior_matches'] = len(row['prior_matches'])
    df['size_prev_match'] = [len(a) for a in df['current_match'].tolist()]
    df['current_match'] = [ [] for _ in range(len(df)) ]
    df['current_group'] = -1

def possible_matches(df):
    # after each match, run this to determine future possible matches
    df['possible_matches'] = [ [] for _ in range(len(df)) ]
    df['num_possible_matches'] = -1
    for index, row in df.iterrows():
        non_previous_matches = df.index.to_numpy()[~np.in1d(df.index.to_numpy(), row['prior_matches'])]
        non_same_department = df.index.to_numpy()[~(row['department'] == df['department'])]
        df.loc[index, 'possible_matches'].extend(np.intersect1d(non_previous_matches, non_same_department))
        df.loc[index, 'num_possible_matches'] = len(df.loc[index, 'possible_matches'])

def perform_random_loop(df):
    # variable to determine when we've succeeded
    out = 0
    # clear any attemps to match that failed
    df['current_match'] = [ [] for _ in range(len(df)) ]
    df['current_group'] = -1
    # create a random column 
    df.loc[:, 'randint'] = np.random.choice(np.arange(0, len(df)), size=len(df), replace=False)
    
    groupnum = 1 # a counter for the group number
    # iterate through, starting with the most number of possible matches 
    for i, (index, row) in enumerate(df.sort_values(['size_prev_match',                                                      'num_possible_matches',                                                      'randint']).iterrows()):
        # select possible matches for person1
        if i == 0:
            p1_possible = df.loc[index, 'possible_matches']
        elif i > 0 :
            if len(remaining) == 0:
                out = 1
                return out
                break
            elif index not in remaining.index.tolist(): 
                continue
            else:
                p1_possible = np.intersect1d(remaining.loc[index, 'possible_matches'],                                              remaining.index.tolist())
        if len(p1_possible) <= 1:
            return out
            break
        # pick a random person2
        p2 = df.loc[p1_possible].sample(1)
        p2_possible = p2['possible_matches'].tolist()
        # take person1 possible matches and remove person2 and all of person2's not possible matches
        p1p2_possible_step1 = np.array(p1_possible)[~np.isin(p1_possible, p2.index.tolist())] # remove p2
        p1p2_possible = p1p2_possible_step1[np.isin(p1p2_possible_step1, p2_possible)]

        if len(p1p2_possible) == 0:
            return out
            break
        # pick a random person3
        p3 = df.loc[p1p2_possible].sample(1)
        p3_possible = p3['possible_matches'].tolist()

        if i < number4groups*4:
            # take person3 out oc p1p2_possible
            p1p2p3_possible_step1 = np.array(p1p2_possible)[~np.isin(p1p2_possible, p3.index.tolist())] 
            # keep only person3's possible matches 
            p1p2p3_possible = p1p2p3_possible_step1[np.isin(p1p2p3_possible_step1, p3_possible)]
            
            if len(p1p2p3_possible) == 0:
                return out
                break
            # pick a random person4
            p4 = df.loc[p1p2p3_possible].sample(1)

            # write the current match for all *4* group members
            df.loc[index, 'current_match'].extend([index, p2.index[0], p3.index[0], p4.index[0]])
            df.loc[p2.index[0], 'current_match'].extend([p2.index[0], index, p3.index[0], p4.index[0]])
            df.loc[p3.index[0], 'current_match'].extend([p3.index[0], index, p2.index[0], p4.index[0]])
            df.loc[p4.index[0], 'current_match'].extend([p4.index[0], index, p2.index[0], p3.index[0]])
            
            df.loc[index, 'current_group'] = groupnum
            df.loc[p2.index[0], 'current_group'] = groupnum
            df.loc[p3.index[0], 'current_group'] = groupnum
            df.loc[p4.index[0], 'current_group'] = groupnum
            
        else:
            # write the current match for all *3* group members
            df.loc[index, 'current_match'].extend([index, p2.index[0], p3.index[0]])
            df.loc[p2.index[0], 'current_match'].extend([p2.index[0], index, p3.index[0]])
            df.loc[p3.index[0], 'current_match'].extend([p3.index[0], index, p2.index[0]])
            
            df.loc[index, 'current_group'] = groupnum
            df.loc[p2.index[0], 'current_group'] = groupnum
            df.loc[p3.index[0], 'current_group'] = groupnum

        # create a new version of the overall df with the matches rows removed
        if i == 0:
            if i < number4groups*4:
                remaining = df.loc[df.index.difference((index, p2.index[0], p3.index[0], p4.index[0]))]
            else:
                remaining = df.loc[df.index.difference((index, p2.index[0], p3.index[0]))]
        if i > 0:
            if i < number4groups*4:
                remaining = remaining.loc[remaining.index.difference((index, p2.index[0],                                                                       p3.index[0], p4.index[0]))]
            else:
                remaining = remaining.loc[remaining.index.difference((index, p2.index[0], p3.index[0]))]

        if i == len(df) - 1:
            out = 1
            return out

        groupnum+=1

def create_match(df):
    # for all matches after the first
    global out
    counter = 0
    out = 0
    while out == 0:
        counter += 1
        out = perform_random_loop(df)  
        print(counter, '\r', end='')
        if counter >= 1000:
            print('match failed, not possible')
            break
    if counter < 1000:
        print('took ',counter, ' tries')
        print('match succeeded')
    return df
    
def write_match(df, pairing_round, filepath):
    global tmp1, tmp2
    # write current match [only] to excel spreadsheet
    # if the first pairing, make a new sheet and copy data
    # second time onward, just add a column
    # pairing_round is ordinal: 1, 2, 3, ... = first round, second round, third round, ...
   
    if pairing_round == 1:
        book = load_workbook(filepath)
        writer = pd.ExcelWriter(filepath, engine = 'openpyxl')
        writer.book = book
        try:
            book.remove(book['pairings'])
        except:
            ''
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df.to_excel(writer, sheet_name='pairings', \
            columns=['first_name', 'last_name', 'email', 'BUID',  'department', 'current_group'],                     header=['first_name', 'last_name', 'email', 'BUID',  'department',                             'group %i'%int(pairing_round)], index=False)
        writer.save()

    elif pairing_round > 1:
        tmp1 = pd.read_excel(filepath, sheet_name='pairings')
        tmp2 = tmp1.copy(deep=True)
        tmp2.insert(len(tmp2.columns), 'group %i'%int(pairing_round), df['current_group'])
        
        book = load_workbook(filepath)
        writer = pd.ExcelWriter(filepath, engine = 'openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        tmp2.to_excel(writer, sheet_name='pairings', index=False)
        writer.save()

def create_match_download_file(data):
    global match_file
    match_file =  data[['first_name', 'last_name', 'email', 'BUID', 'department', 'current_group']]
    return match_file

def create_download_link(df, title="Download CSV file", filename="data.csv"):  
    csv = df.to_csv()
    b64 = base64.b64encode(csv.encode())
    payload = b64.decode()
    html = '<a download="{filename}" href="data:text/csv;base64,{payload}" target="_blank">{title}</a>'
    html = html.format(payload=payload,title=title,filename=filename)
    display(HTML(html))

style = {'description_width': 'initial'}

file_widget = widgets.Text(
    description='Participants file:', 
    disabled=False, 
    style=style)

round_widget = widgets.Dropdown(
    options=np.arange(1, 10),
    description='Matching round:', 
    disabled=False, 
    style=style)
import_button = widgets.Button(description='Import', style=style)
#possible_matches_botton = widgets.Button(description='Possible Matches', style=style)
create_match_button = widgets.Button(description='Create Match', style=style)
#write_match_button = widgets.Button(description='Write Match', style=style)
complete_match_button = widgets.Button(description='Complete Match', style=style)
event_download_match_button = widgets.Button(description='Download Match', style=style)
event_download_data_button = widgets.Button(description='Download Data', style=style)

def display_widget():
    display(file_widget), \
    display(import_button), \
    display(round_widget), \
#    display(possible_matches_botton), \
    display(create_match_button), \
#    display(write_match_button), \
    display(event_download_match_button), \
    display(complete_match_button), \
    display(event_download_data_button)

data = None
number3groups = None
number4groups = None

def event_import(button):
    global data, number3groups, number4groups
    clear_output()
    display_widget()
    data = pd.read_excel(file_widget.value)
    init(data)
    
    n = len(data)
    number3groups = n // 3
    number4groups = n - (number3groups * 3)

    print('imported participants')

def event_create_match(button):
    global data
    clear_output()
    display_widget()
    possible_matches(data)
    print('round %i possible matches complete'%round_widget.value)

    data = create_match(data)
    if out == 1:
        print('round %i match created'%round_widget.value)
    if out == 0:
        print('round %i match failed'%round_widget.value)

def event_write_match(button):
    global data
    clear_output()
    display_widget()
    write_match(data, round_widget.value, file_widget.value)
    print('round %i match written'%round_widget.value)

def event_complete_match(button):
    global data
    clear_output()
    display_widget()
    complete_match(data)
    print('round %i match complete'%round_widget.value)

def event_download_match(button):
    clear_output()
    display_widget()
    match_file = create_match_download_file(data)
    return create_download_link(match_file, title="download round %i match"%round_widget.value, \
                         filename="round_%i_match.csv"%round_widget.value)

def event_download_data(button):
    clear_output()
    display_widget()
    return create_download_link(data, title="download round %i data"%round_widget.value, \
                         filename="round_%i_data.csv"%round_widget.value)

import_button.on_click(event_import)
create_match_button.on_click(event_create_match)
complete_match_button.on_click(event_complete_match)
event_download_match_button.on_click(event_download_match)
event_download_data_button.on_click(event_download_data)


